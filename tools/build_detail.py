import json
import re
from pathlib import Path
from openpyxl import load_workbook
import urllib.parse

UPLOAD_DIR = Path("uploads")
OUT_BASE = Path("site") / "snapshots"  # gh-pages는 site를 배포하니까 여기로!

def pick_date_key(stem: str) -> str:
    m = re.search(r"(\d{4}_\d{2}_\d{2})", stem)
    if not m:
        return ""
    return m.group(1)

def normalize_header(v):
    if v is None:
        return ""
    return str(v).strip().replace("\n", "").replace("\r", "").replace(" ", "")

def build_header_map(ws, header_row=1, max_cols=120):
    """
    서버 시트에서 필요한 컬럼(결사명/클래스/토벌등급)을 헤더로 찾아 컬럼 번호 매핑
    """
    m = {}
    for col in range(1, max_cols + 1):
        key = normalize_header(ws.cell(row=header_row, column=col).value)

        if key in ("결사명", "결사"):
            m["guild"] = col
        elif key in ("클래스", "직업"):
            m["class"] = col
        elif key in ("토벌등급", "등급"):
            m["grade"] = col
    return m

def is_server_sheet(ws) -> bool:
    h = build_header_map(ws)
    return all(k in h for k in ("guild", "class", "grade"))

def safe_str(v):
    return "" if v is None else str(v).strip()

def safe_int(v):
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if s == "":
        return 0
    try:
        return int(float(s))
    except:
        return 0

def build_detail_for_xlsx(xlsx_path: Path):
    date_key = pick_date_key(xlsx_path.stem)
    if not date_key:
        print(f"skip (no date in filename): {xlsx_path.name}")
        return

    out_dir = OUT_BASE / f"detail_{date_key}"
    out_dir.mkdir(parents=True, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=True)

    made = 0
    for name in wb.sheetnames:
        ws = wb[name]

        if not is_server_sheet(ws):
            continue

        header = build_header_map(ws)
        gcol, ccol, grcol = header["guild"], header["class"], header["grade"]

        guilds = {}  # guild -> {members, byClass, byGrade}

        for r in range(2, ws.max_row + 1):
            guild = safe_str(ws.cell(row=r, column=gcol).value)
            if guild == "":
                continue

            cls = safe_str(ws.cell(row=r, column=ccol).value)
            grade = safe_int(ws.cell(row=r, column=grcol).value)

            g = guilds.get(guild)
            if not g:
                g = {"members": 0, "byClass": {}, "byGrade": {}}
                guilds[guild] = g

            g["members"] += 1

            if cls == "":
                cls = "미확인"
            g["byClass"][cls] = int(g["byClass"].get(cls, 0) + 1)

            # 등급은 문자열 키로 저장(웹에서 그대로 표기)
            sk = str(grade) if grade else "0"
            g["byGrade"][sk] = int(g["byGrade"].get(sk, 0) + 1)

        out = {
            "dateKey": date_key,
            "server": name,
            "guilds": guilds
        }

        # 파일명은 encodeURIComponent(server)와 동일하게 URL 인코딩
        fname = urllib.parse.quote(name, safe="") + ".json"
        (out_dir / fname).write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        made += 1

    print(f"[detail] {xlsx_path.name} -> {out_dir} (server sheets: {made})")

def main():
    OUT_BASE.mkdir(parents=True, exist_ok=True)

    files = sorted(UPLOAD_DIR.glob("ranking_????_??_??.xlsx"))
    if not files:
        print("no xlsx in uploads/")
        return

    for f in files:
        build_detail_for_xlsx(f)

if __name__ == "__main__":
    main()
