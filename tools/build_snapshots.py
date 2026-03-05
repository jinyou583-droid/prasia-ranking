import os
import json
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

# 입력/출력 경로
UPLOAD_DIR = Path("uploads")
OUT_DIR = Path("site") / "snapshots"

OUT_DIR.mkdir(parents=True, exist_ok=True)

def safe_str(x):
    if x is None:
        return ""
    return str(x).strip()

def guess_label_from_filename(stem: str) -> str:
    # 예: ranking_2026_03_05 -> 2026-03-05
    parts = stem.replace("-", "_").split("_")
    nums = [p for p in parts if p.isdigit()]
    if len(nums) >= 3:
        y, m, d = nums[-3], nums[-2], nums[-1]
        if len(y) == 4:
            return f"{y}-{m.zfill(2)}-{d.zfill(2)}"
    return stem

def parse_sheet(ws):
    """
    기대 컬럼(왼쪽부터):
    A: rank
    B: guild
    C: server
    D: hunt_score
    E: level_score
    F: total_score
    """
    rows = []
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r:
            continue

        # r 길이가 짧을 수 있어서 안전하게 꺼냄
        rank = r[0] if len(r) > 0 else None
        guild = r[1] if len(r) > 1 else None
        server = r[2] if len(r) > 2 else None
        hunt_score = r[3] if len(r) > 3 else None
        level_score = r[4] if len(r) > 4 else None
        total_score = r[5] if len(r) > 5 else None

        # rank / guild / server 중 핵심이 비면 스킵
        if rank is None and not guild and not server:
            continue

        rows.append({
            "rank": rank,
            "guild": safe_str(guild),
            "server": safe_str(server),
            "hunt_score": hunt_score,
            "level_score": level_score,
            "total_score": total_score
        })
    return rows

index = []

xlsx_files = sorted(UPLOAD_DIR.glob("*.xlsx"))
for xlsx_path in xlsx_files:
    stem = xlsx_path.stem
    label = guess_label_from_filename(stem)
    out_name = f"{stem}.json"
    out_path = OUT_DIR / out_name

    wb = load_workbook(xlsx_path, data_only=True)
    # 첫 시트 사용 (원하면 여기서 특정 시트명으로 바꿀 수 있음)
    ws = wb[wb.sheetnames[0]]

    data = parse_sheet(ws)

    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    index.append({
        "label": label,
        "file": out_name,
        "rows": len(data)
    })

# 날짜 느낌으로 정렬(라벨이 YYYY-MM-DD면 최신이 위로)
def sort_key(item):
    try:
        return datetime.strptime(item["label"], "%Y-%m-%d")
    except Exception:
        return datetime.min

index.sort(key=sort_key, reverse=True)

# index.json 생성
with open(OUT_DIR / "index.json", "w", encoding="utf-8") as f:
    json.dump(index, f, ensure_ascii=False, indent=2)

print(f"Generated {len(index)} snapshot(s) into {OUT_DIR}")
